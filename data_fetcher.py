import requests
import os
import sys
import pandas as pd
from dotenv import load_dotenv
from datetime import datetime, timedelta
import random

load_dotenv()

class FootballDataFetcher:
    def __init__(self):
        self.api_key = os.getenv("FOOTBALL_DATA_API_KEY")
        self.apisports_key = os.getenv("APISPORTS_API_KEY")
        self.base_url = "https://api.football-data.org/v4"
        self.apisports_url = "https://v3.football.api-sports.io"
        self.headers = {"X-Auth-Token": self.api_key} if self.api_key else {}
        self.apisports_headers = {"x-apisports-key": self.apisports_key} if self.apisports_key else {}

        # ── MASTER TEAM ID MAP ──────────────────────────────────
        self.master_teams = {
            # LaLiga
            "Real Madrid CF": 86, "FC Barcelona": 81, "Atlético de Madrid": 78,
            "Villarreal CF": 299, "Real Betis": 90, "RCD Espanyol": 80,
            "RC Celta": 82, "Real Sociedad": 92, "Athletic Club": 77,
            "CA Osasuna": 79, "Getafe CF": 88, "Sevilla FC": 559,
            "Deportivo Alavés": 263, "Valencia CF": 95, "Girona FC": 298,
            "Rayo Vallecano": 87, "RCD Mallorca": 89, "Levante UD": 264,
            "Real Valladolid": 250, "UD Las Palmas": 275, "CD Leganés": 345,
            # Premier League
            "Arsenal FC": 57, "Manchester City FC": 65, "Liverpool FC": 64,
            "Aston Villa FC": 58, "Manchester United FC": 66, "Chelsea FC": 61,
            "Tottenham Hotspur FC": 73, "Newcastle United FC": 67,
            "West Ham United FC": 563, "Brighton & Hove Albion FC": 397,
            "Everton FC": 62, "Fulham FC": 63, "Bournemouth": 1044,
            "Crystal Palace FC": 354, "Wolverhampton Wanderers": 76,
            "Leicester City FC": 338, "Brentford FC": 402, "Nottingham Forest FC": 351,
            # Serie A
            "Inter Milan": 108, "AC Milan": 98, "Juventus FC": 109,
            "SSC Napoli": 113, "AS Roma": 100, "Atalanta": 102, "SS Lazio": 110,
            "Bologna": 103, "ACF Fiorentina": 99, "Torino FC": 586,
            "Genoa CFC": 107, "Monza": 586, "Como 1907": 5890,
            # Bundesliga
            "FC Bayern München": 5, "Bayer 04 Leverkusen": 3, "Borussia Dortmund": 4,
            "RB Leipzig": 721, "VfB Stuttgart": 10, "Eintracht Frankfurt": 9,
            "SC Freiburg": 17, "TSG Hoffenheim": 533, "Werder Bremen": 12,
            "Union Berlin": 28513, "Borussia Mönchengladbach": 18,
            "FC Augsburg": 16, "VfL Wolfsburg": 11,
            # Ligue 1
            "Paris Saint-Germain FC": 524, "AS Monaco FC": 548,
            "Olympique de Marseille": 516, "Lille OSC": 521,
            "Olympique Lyonnais": 523, "OGC Nice": 522, "Stade Rennais FC": 529,
            "RC Lens": 532, "Stade de Reims": 547,
            # Serie B
            "Sassuolo": 105, "Sampdoria": 103, "Palermo": 115,
            # Argentina
            "River Plate": 3211, "Boca Juniors": 3201, "Vélez Sarsfield": 3215,
            "Racing Club": 3203, "Independiente": 3207, "San Lorenzo": 3214,
            "Estudiantes L.P.": 3204, "Rosario Central": 3213,
            "Newell's Old Boys": 3209, "Talleres": 3217, "Huracán": 3205,
            "Belgrano": 3202, "Defensa y Justicia": 3220, "Lanús": 3208,
            # Brasil
            "Flamengo": 5981, "Palmeiras": 1963, "Botafogo": 1967,
            "Corinthians": 1977, "São Paulo FC": 5982, "Fluminense": 1961,
            "Internacional": 1974, "Grêmio": 1972, "Atlético Mineiro": 1966,
            "Vasco da Gama": 1980, "Bahia": 1968, "Cruzeiro": 1971,
            "Fortaleza": 3573, "Santos FC": 1978,
            # Colombia
            "Atlético Nacional": 2274, "Millonarios": 2283, "Santa Fe": 2285,
            "Junior": 2279, "América de Cali": 2272, "Deportivo Cali": 2276,
            "Deportes Tolima": 2277, "Once Caldas": 2284,
            # México
            "Cruz Azul": 1951, "Tigres UANL": 1958, "Club América": 1946,
            "Chivas Guadalajara": 1953, "Monterrey": 1955, "Pumas UNAM": 1956,
            "Toluca": 1959, "Atlas": 1949, "León": 1954, "Tijuana": 1957,
            "Atlético San Luis": 10255, "Mazatlán": 2282,
            # Saudi Pro League
            "Al Hilal": 2822, "Al Ittihad": 2814, "Al Nassr": 2818,
            "Al Ahli": 2808, "Al Shabab": 2823, "Al Ettifaq": 2811,
            # MLS
            "Inter Miami CF": 306781, "LA Galaxy": 351, "LAFC": 31277,
            "Columbus Crew": 326, "Real Salt Lake": 1903, "Orlando City SC": 8586,
            # Other
            "Sporting CP": 498, "FC Porto": 503, "SL Benfica": 294,
            "PSV Eindhoven": 397, "Ajax": 678, "Feyenoord": 675,
            "Galatasaray": 442, "Fenerbahçe": 441, "Beşiktaş": 440,
        }

        # ── TEAM DATABASE (stadium, coach, country) ──────────────
        self.team_database = {
            "Real Madrid CF": {"stadium": "Santiago Bernabéu", "coach": "Carlo Ancelotti", "country": "España", "league": "LaLiga"},
            "FC Barcelona": {"stadium": "Spotify Camp Nou", "coach": "Hansi Flick", "country": "España", "league": "LaLiga"},
            "Atlético de Madrid": {"stadium": "Cívitas Metropolitano", "coach": "Diego Simeone", "country": "España", "league": "LaLiga"},
            "Manchester City FC": {"stadium": "Etihad Stadium", "coach": "Pep Guardiola", "country": "Inglaterra", "league": "Premier League"},
            "Liverpool FC": {"stadium": "Anfield", "coach": "Arne Slot", "country": "Inglaterra", "league": "Premier League"},
            "Arsenal FC": {"stadium": "Emirates Stadium", "coach": "Mikel Arteta", "country": "Inglaterra", "league": "Premier League"},
            "Inter Milan": {"stadium": "San Siro", "coach": "Simone Inzaghi", "country": "Italia", "league": "Serie A"},
            "FC Bayern München": {"stadium": "Allianz Arena", "coach": "Vincent Kompany", "country": "Alemania", "league": "Bundesliga"},
            "Paris Saint-Germain FC": {"stadium": "Parc des Princes", "coach": "Luis Enrique", "country": "Francia", "league": "Ligue 1"},
            "River Plate": {"stadium": "Más Monumental", "coach": "Marcelo Gallardo", "country": "Argentina", "league": "Liga Profesional"},
            "Boca Juniors": {"stadium": "La Bombonera", "coach": "Fernando Gago", "country": "Argentina", "league": "Liga Profesional"},
            "Flamengo": {"stadium": "Maracanã", "coach": "Filipe Luís", "country": "Brasil", "league": "Brasileirão"},
            "Palmeiras": {"stadium": "Allianz Parque", "coach": "Abel Ferreira", "country": "Brasil", "league": "Brasileirão"},
        }

    def get_crest(self, name):
        tid = self.master_teams.get(name)
        if not tid:
            for n, i in self.master_teams.items():
                if name.lower() in n.lower() or n.lower() in name.lower():
                    tid = i
                    break
        if tid and tid < 3000:
            return f"https://crests.football-data.org/{tid}.png"
        initials = "".join([w[0] for w in name.split()[:2]]).upper()
        return f"https://ui-avatars.com/api/?name={initials}&background=0d1117&color=00e87a&bold=true&size=40"

    def _get_complete_table_data(self):
        return {
            # ── COPAS INTERNACIONALES ────────────────────
            "UCL": ["Real Madrid CF", "FC Barcelona", "Manchester City FC", "Arsenal FC", "Liverpool FC", "Inter Milan", "AC Milan", "Juventus FC", "FC Bayern München", "Bayer 04 Leverkusen", "Borussia Dortmund", "Paris Saint-Germain FC", "Atlético de Madrid", "Sporting CP", "SL Benfica", "Feyenoord", "PSV Eindhoven", "Celtic FC", "Aston Villa FC", "Bologna", "Atalanta", "AS Monaco FC", "Lille OSC", "Stade Brestois 29"],
            "LIB": ["River Plate", "Boca Juniors", "Flamengo", "Palmeiras", "Fluminense", "Botafogo", "Atlético Mineiro", "São Paulo FC", "Peñarol", "Club Nacional de Football", "Club Olimpia", "Cerro Porteño", "Colo-Colo", "Barcelona SC", "LDU Quito", "Junior", "Millonarios", "Estudiantes L.P.", "Libertad", "Bolívar", "The Strongest", "Talleres", "San Lorenzo", "Independiente del Valle"],
            
            # ── EUROPA ──────────────────────────────────
            "PD": ["Real Madrid CF", "FC Barcelona", "Atlético de Madrid", "Villarreal CF", "Real Betis",
                   "Real Sociedad", "Athletic Club", "Girona FC", "CA Osasuna", "Sevilla FC",
                   "RCD Espanyol", "RC Celta", "Valencia CF", "Getafe CF", "Deportivo Alavés",
                   "Rayo Vallecano", "RCD Mallorca", "Levante UD", "Real Valladolid", "UD Las Palmas"],
            "ELC": ["Leeds United", "Burnley FC", "Sheffield United", "Sunderland AFC", "West Bromwich Albion", "Middlesbrough FC", "Coventry City", "Norwich City", "Luton Town", "Watford FC", "Blackburn Rovers", "Derby County", "Stoke City", "Hull City", "Swansea City", "Preston North End"],
            "PL": ["Manchester City FC", "Arsenal FC", "Liverpool FC", "Aston Villa FC", "Chelsea FC",
                   "Tottenham Hotspur FC", "Manchester United FC", "Newcastle United FC",
                   "West Ham United FC", "Brighton & Hove Albion FC", "Everton FC", "Fulham FC",
                   "Bournemouth", "Crystal Palace FC", "Wolverhampton Wanderers",
                   "Leicester City FC", "Brentford FC", "Nottingham Forest FC"],
            "SA": ["Inter Milan", "AC Milan", "Juventus FC", "SSC Napoli", "AS Roma", "Atalanta",
                   "SS Lazio", "Bologna", "ACF Fiorentina", "Torino FC", "Genoa CFC",
                   "Monza", "Como 1907", "Hellas Verona FC", "Cagliari Calcio",
                   "US Lecce", "Parma Calcio 1913", "Venezia FC", "Empoli FC", "Udinese Calcio"],
            "BL1": ["FC Bayern München", "Bayer 04 Leverkusen", "Borussia Dortmund", "RB Leipzig",
                    "VfB Stuttgart", "Eintracht Frankfurt", "SC Freiburg", "TSG Hoffenheim",
                    "Werder Bremen", "Union Berlin", "Borussia Mönchengladbach", "FC Augsburg",
                    "VfL Wolfsburg", "1. FC Heidenheim 1846", "FC St. Pauli", "1. FSV Mainz 05"],
            "FL1": ["Paris Saint-Germain FC", "AS Monaco FC", "Olympique de Marseille", "Lille OSC",
                    "Olympique Lyonnais", "OGC Nice", "Stade Rennais FC", "RC Lens",
                    "Stade de Reims", "RC Strasbourg Alsace", "Stade Brestois 29", "Montpellier HSC",
                    "Le Havre AC", "FC Nantes", "Toulouse FC", "Angers SCO"],
            "PPL": ["Sporting CP", "FC Porto", "SL Benfica", "SC Braga", "Vitória SC",
                    "Moreirense FC", "Famalicão", "Santa Clara", "Boavista FC", "Casa Pia AC",
                    "Gil Vicente FC", "Arouca FC", "FC Vizela", "Estoril Praia"],
            "DED": ["PSV Eindhoven", "Ajax", "Feyenoord", "AZ Alkmaar", "FC Twente",
                    "FC Utrecht", "Go Ahead Eagles", "NEC Nijmegen", "Sparta Rotterdam",
                    "FC Groningen", "SC Heerenveen", "PEC Zwolle", "Almere City FC", "NAC Breda"],
            "TSL": ["Galatasaray", "Fenerbahçe", "Beşiktaş", "Trabzonspor", "Başakşehir",
                    "Samsunspor", "Eyüpspor", "Konyaspor", "Sivasspor", "Antalyaspor",
                    "Kasımpaşa", "Rizespor", "Gaziantep FK", "Alanyaspor"],
            "BEL": ["RSC Anderlecht", "Club Brugge KV", "K.A.A. Gent", "Royal Antwerp FC",
                    "KRC Genk", "Union Saint-Gilloise", "Standard Liège", "KV Mechelen",
                    "Cercle Brugge KSV", "OH Leuven", "Westerlo", "Beerschot VA"],
            "SPD": ["Real Zaragoza", "UD Almería", "Racing Santander", "CD Castellón",
                    "Sporting Gijón", "Málaga CF", "SD Eibar", "Real Oviedo",
                    "CD Leganés", "Granada CF", "Levante UD", "Tenerife",
                    "Eldense", "Ferrol", "Mirandes", "Albacete"],
            "SB": ["Sassuolo", "Sampdoria", "Palermo", "Cremonese", "Pisa SC",
                   "Spezia Calcio", "Modena FC", "Brescia Calcio", "Bari",
                   "Frosinone Calcio", "Cosenza Calcio", "Catanzaro"],
            "SPR": ["Zenit St. Petersburg", "CSKA Moscow", "Spartak Moscow", "Lokomotiv Moscow",
                    "Dynamo Moscow", "Krasnodar", "Akhmat Grozny", "Rubin Kazan"],
            "SPO": ["Celtic FC", "Rangers FC", "Heart of Midlothian FC", "Hibernian FC",
                    "Aberdeen FC", "St. Mirren FC", "Dundee United FC", "Motherwell FC"],
            "AUS": ["Red Bull Salzburg", "SK Sturm Graz", "FK Austria Wien",
                    "Rapid Wien", "RB Salzburg", "LASK"],
            # ── AMÉRICAS ────────────────────────────────
            "ARG": ["Vélez Sarsfield", "River Plate", "Racing Club", "Talleres", "Huracán",
                    "Boca Juniors", "Independiente", "Estudiantes L.P.", "Lanús", "Godoy Cruz",
                    "Unión", "Belgrano", "Instituto", "Argentinos Juniors", "Defensa y Justicia",
                    "Rosario Central", "Platense", "San Lorenzo", "Banfield", "Gimnasia L.P.",
                    "Tigre", "Barracas Central", "Newell's Old Boys", "Tucumán",
                    "Ind. Rivadavia", "Sarmiento", "Central Córdoba", "Riestra"],
            "BRA": ["Botafogo", "Palmeiras", "Flamengo", "Fortaleza", "Internacional",
                    "São Paulo FC", "Bahia", "Cruzeiro", "Atlético Mineiro", "Vasco da Gama",
                    "Corinthians", "Grêmio", "Fluminense", "Santos FC", "Bragantino",
                    "Athletico Paranaense", "Goiás EC", "Cuiabá", "América Mineiro", "Criciúma"],
            "COL": ["Deportivo Cali", "Atlético Nacional", "Millonarios", "Santa Fe", "Junior",
                    "América de Cali", "Once Caldas", "Deportes Tolima", "Peñarol",
                    "Ind. Medellín", "Envigado FC", "Bucaramanga"],
            "MX": ["Cruz Azul", "Tigres UANL", "Toluca", "Pumas UNAM", "Monterrey",
                   "Club América", "Atlético San Luis", "Chivas Guadalajara", "Tijuana",
                   "Atlas", "León", "Necaxa", "Puebla", "Juárez", "Mazatlán",
                   "Pachuca", "Santos Laguna", "FC Juárez"],
            "MLS": ["Inter Miami CF", "Columbus Crew", "LA Galaxy", "LAFC", "FC Cincinnati",
                    "Real Salt Lake", "Orlando City SC", "Portland Timbers", "Seattle Sounders",
                    "New York City FC", "Atlanta United FC", "Nashville SC",
                    "New England Revolution", "Colorado Rapids", "Austin FC"],
            "ECU": ["LDU Quito", "Emelec", "Barcelona SC", "Independiente del Valle",
                    "El Nacional", "Aucas", "Delfín", "Orense"],
            "CHI": ["Club Universidad de Chile", "Colo-Colo", "Universidad Católica",
                    "Deportes Antofagasta", "Huachipato", "Audax Italiano"],
            "URU": ["Peñarol", "Club Nacional de Football", "Defensor Sporting",
                    "Danubio", "River Plate Montevideo"],
            "PAR": ["Club Olimpia", "Cerro Porteño", "Libertad", "Guaraní",
                    "Sol de América", "Deportivo Luqueño"],
            "PER": ["Universitario", "Alianza Lima", "Sporting Cristal", "Melgar", "Cienciano", "Cusco FC", "ADT Tarma", "Sport Boys", "César Vallejo", "Carlos A. Mannucci", "UTC de Cajamarca", "Los Chankas", "Atlético Grau", "Comerciantes Unidos"],
            "VEN": ["Deportivo Táchira", "Caracas FC", "Academia Puerto Cabello", "Monagas SC", "Metropolitanos FC", "Estudiantes de Mérida", "Deportivo La Guaira", "Portuguesa FC", "Angostura FC", "Carabobo FC", "Zamora FC", "Universidad Central"],
            "BOL": ["Bolívar", "The Strongest", "Always Ready", "Blooming", "Oriente Petrolero", "Jorge Wilstermann", "Nacional Potosí", "Real Tomayapo", "San José", "Aurora", "Universitario de Vinto", "Guabirá"],
            # ── ASIA Y OTROS ────────────────────────────
            "SPL": ["Al Hilal", "Al Ittihad", "Al Nassr", "Al Ahli", "Al Shabab",
                    "Al Ettifaq", "Al Taawoun", "Al Qadsiah", "Al Fayha", "Damac FC"],
            "J1": ["Vissel Kobe", "Sanfrecce Hiroshima", "Machida Zelvia", "Gamba Osaka",
                   "Kashima Antlers", "Kawasaki Frontale", "Yokohama F. Marinos",
                   "Cerezo Osaka", "Nagoya Grampus", "Urawa Red Diamonds"],
            "CSL": ["Shanghai Port", "Shanghai Shenhua", "Chengdu Rongcheng",
                    "Beijing Guoan", "Wuhan Three Towns", "Shandong Taishan",
                    "Zhejiang FC", "Tianjin Jinmen Tiger"],
            "ISL": ["Mumbai City FC", "Mohun Bagan SG", "Bengaluru FC", "Kerala Blasters",
                    "FC Goa", "ATK Mohun Bagan", "Hyderabad FC", "Odisha FC",
                    "Jamshedpur FC", "Northeast United FC"],
            "CAF": ["Ahly SC", "Zamalek SC", "Pyramids FC", "Esperance Sportive",
                    "Mamelodi Sundowns", "Kaizer Chiefs", "Al-Merrikh", "Wydad Athletic"],
            "NIG": ["Enyimba FC", "Shooting Stars SC", "Rivers United", "Kano Pillars",
                    "Heartland FC", "Lobi Stars", "Plateau United"],
            "GRE": ["AEK Athens", "Olympiacos", "PAOK Salonika", "Panathinaikos", "Aris", "Asteras Tripolis", "Atromitos", "OFI Crete", "Lamia", "Panetolikos", "Volos", "Panserraikos"],
        }

    # ── WORLD CUP 2026 DATA ──────────────────────────────────────
    def get_worldcup_data(self):
        """Returns FIFA World Cup 2026 group data with standings and matches."""
        # Grupos OFICIALES FIFA World Cup 2026 – 48 equipos confirmados
        draw = {
            "A": ["México", "Sudáfrica", "Corea del Sur", "Chequia"],
            "B": ["Canadá", "Bosnia", "Qatar", "Suiza"],
            "C": ["Brasil", "Marruecos", "Haití", "Escocia"],
            "D": ["Estados Unidos", "Paraguay", "Australia", "Turquía"],
            "E": ["Alemania", "Curazao", "Costa de Marfil", "Ecuador"],
            "F": ["Países Bajos", "Japón", "Ucrania", "Túnez"],
            "G": ["Bélgica", "Egipto", "Irán", "Nueva Zelanda"],
            "H": ["España", "Cabo Verde", "Arabia Saudí", "Uruguay"],
            "I": ["Francia", "Senegal", "Bolivia", "Suecia"],
            "J": ["Argentina", "Argelia", "Austria", "Jordania"],
            "K": ["Portugal", "Rep. D. Congo", "Uzbekistán", "Colombia"],
            "L": ["Inglaterra", "Croacia", "Ghana", "Panamá"]
        }
        
        flags = {
            # Grupo A
            "México": "🇲🇽", "Sudáfrica": "🇿🇦", "Corea del Sur": "🇰🇷", "Chequia": "🇨🇿",
            # Grupo B
            "Canadá": "🇨🇦", "Bosnia": "🇧🇦", "Qatar": "🇶🇦", "Suiza": "🇨🇭",
            # Grupo C
            "Brasil": "🇧🇷", "Marruecos": "🇲🇦", "Haití": "🇭🇹", "Escocia": "🏴󠁧󠁢󠁳󠁣󠁴󠁿",
            # Grupo D
            "Estados Unidos": "🇺🇸", "Paraguay": "🇵🇾", "Australia": "🇦🇺", "Turquía": "🇹🇷",
            # Grupo E
            "Alemania": "🇩🇪", "Curazao": "🇨🇼", "Costa de Marfil": "🇨🇮", "Ecuador": "🇪🇨",
            # Grupo F
            "Países Bajos": "🇳🇱", "Japón": "🇯🇵", "Ucrania": "🇺🇦", "Túnez": "🇹🇳",
            # Grupo G
            "Bélgica": "🇧🇪", "Egipto": "🇪🇬", "Irán": "🇮🇷", "Nueva Zelanda": "🇳🇿",
            # Grupo H
            "España": "🇪🇸", "Cabo Verde": "🇨🇻", "Arabia Saudí": "🇸🇦", "Uruguay": "🇺🇾",
            # Grupo I
            "Francia": "🇫🇷", "Senegal": "🇸🇳", "Bolivia": "🇧🇴", "Suecia": "🇸🇪",
            # Grupo J
            "Argentina": "🇦🇷", "Argelia": "🇩🇿", "Austria": "🇦🇹", "Jordania": "🇯🇴",
            # Grupo K – Congo DR confirmado
            "Portugal": "🇵🇹", "Rep. D. Congo": "🇨🇩", "Uzbekistán": "🇺🇿", "Colombia": "🇨🇴",
            # Grupo L
            "Inglaterra": "🏴󠁧󠁢󠁥󠁮󠁧󠁿", "Croacia": "🇭🇷", "Ghana": "🇬🇭", "Panamá": "🇵🇦"
        }

        groups = {}
        
        # J3 dates and schedule mapping
        dates = {
            "A": ("11 Jun", "15 Jun", "19 Jun"),
            "B": ("11 Jun", "15 Jun", "19 Jun"),
            "C": ("12 Jun", "16 Jun", "20 Jun"),
            "D": ("12 Jun", "16 Jun", "20 Jun"),
            "E": ("12 Jun", "16 Jun", "21 Jun"),
            "F": ("13 Jun", "17 Jun", "21 Jun"),
            "G": ("13 Jun", "17 Jun", "21 Jun"),
            "H": ("13 Jun", "17 Jun", "22 Jun"),
            "I": ("14 Jun", "18 Jun", "22 Jun"),
            "J": ("14 Jun", "18 Jun", "22 Jun"),
            "K": ("14 Jun", "18 Jun", "23 Jun"),
            "L": ("14 Jun", "18 Jun", "23 Jun"),
        }
        
        # Precooked realistic results for (M1, M2, M3, M4) per group
        # Each group gets slightly different but realistic results to make the standings interesting
        precooked_scores = {
            # (J1-M1, J1-M2, J2-M1, J2-M2)
            "A": ("2-1", "1-1", "2-0", "1-1"), # Mex-RSA 2-1, Kor-Cze 1-1, Mex-Kor 2-0, RSA-Cze 1-1
            "B": ("1-1", "1-2", "6-0", "0-2"), # Can-Bos 1-1, Qat-Sui 1-2, Can-Qat 6-0, Bos-Sui 0-2
            "C": ("3-1", "1-2", "2-0", "1-1"), # Bra-Mar 3-1, Hai-Sco 1-2, Bra-Hai 2-0, Mar-Sco 1-1
            "D": ("2-1", "1-2", "3-1", "0-2"), # USA-Par 2-1, Aus-Tur 1-2, USA-Aus 3-1, Par-Tur 0-2
            "E": ("2-0", "1-2", "1-1", "1-3"), # Ger-Cur 2-0, CIV-Ecu 1-2, Ger-CIV 1-1, Cur-Ecu 1-3
            "F": ("2-1", "2-0", "3-1", "1-1"), # Ned-Jap 2-1, Ukr-Tun 2-0, Ned-Ukr 3-1, Jap-Tun 1-1
            "G": ("1-0", "0-0", "2-1", "1-1"), # Bel-Egy 1-0, Ira-NZL 0-0, Bel-Ira 2-1, Egy-NZL 1-1
            "H": ("2-0", "1-3", "2-1", "0-2"), # Esp-CPV 2-0, Sau-Uru 1-3, Esp-Sau 2-1, CPV-Uru 0-2
            "I": ("3-0", "1-2", "2-1", "1-1"), # Fra-Sen 3-0, Bol-Nor 1-2, Fra-Bol 2-1, Sen-Nor 1-1
            "J": ("2-0", "1-1", "3-1", "1-2"), # Arg-Alg 2-0, Aut-Jor 1-1, Arg-Aut 3-1, Alg-Jor 1-2
            "K": ("1-1", "1-2", "2-0", "1-0"), # Por-Congo 1-1, Uzb-Col 1-2, Por-Uzb 2-0, Congo-Col 1-0
            "L": ("2-1", "1-1", "2-0", "1-2"), # Eng-Cro 2-1, Gha-Pan 1-1, Eng-Gha 2-0, Cro-Pan 1-2
        }

        for letter, teams_list in draw.items():
            g_dates = dates[letter]
            scores = precooked_scores[letter]
            
            # Setup teams structures
            g_teams = []
            for name in teams_list:
                g_teams.append({
                    "name": name,
                    "flag": flags.get(name, "🏳️"),
                    "pj": 0, "v": 0, "e": 0, "d": 0,
                    "gf": 0, "gc": 0, "dg": 0, "pts": 0
                })
                
            # Setup matches structures
            t0, t1, t2, t3 = teams_list
            g_matches = [
                # J1
                {"home": t0, "away": t1, "hflag": flags[t0], "aflag": flags[t1], "score": scores[0], "date": g_dates[0], "status": "FT"},
                {"home": t2, "away": t3, "hflag": flags[t2], "aflag": flags[t3], "score": scores[1], "date": g_dates[0], "status": "FT"},
                # J2
                {"home": t0, "away": t2, "hflag": flags[t0], "aflag": flags[t2], "score": scores[2], "date": g_dates[1], "status": "FT"},
                {"home": t1, "away": t3, "hflag": flags[t1], "aflag": flags[t3], "score": scores[3], "date": g_dates[1], "status": "FT"},
                # J3 (PR)
                {"home": t3, "away": t0, "hflag": flags[t3], "aflag": flags[t0], "score": "-", "date": g_dates[2], "status": "PR"},
                {"home": t1, "away": t2, "hflag": flags[t1], "aflag": flags[t2], "score": "-", "date": g_dates[2], "status": "PR"},
            ]
            
            # Calculate standings for J1 & J2
            for m in g_matches:
                if m["status"] == "FT":
                    h_goals, a_goals = map(int, m["score"].split("-"))
                    home_team = next(t for t in g_teams if t["name"] == m["home"])
                    away_team = next(t for t in g_teams if t["name"] == m["away"])
                    
                    home_team["pj"] += 1
                    away_team["pj"] += 1
                    home_team["gf"] += h_goals
                    home_team["gc"] += a_goals
                    away_team["gf"] += a_goals
                    away_team["gc"] += h_goals
                    
                    if h_goals > a_goals:
                        home_team["v"] += 1
                        home_team["pts"] += 3
                        away_team["d"] += 1
                    elif a_goals > h_goals:
                        away_team["v"] += 1
                        away_team["pts"] += 3
                        home_team["d"] += 1
                    else:
                        home_team["e"] += 1
                        home_team["pts"] += 1
                        away_team["e"] += 1
                        away_team["pts"] += 1
            
            for t in g_teams:
                t["dg"] = t["gf"] - t["gc"]
                
            # Sort: Pts -> DG -> GF
            g_teams.sort(key=lambda x: (x["pts"], x["dg"], x["gf"]), reverse=True)
            
            groups[letter] = {
                "teams": g_teams,
                "matches": g_matches
            }
            
        return groups
        return groups

    def get_standings(self, league="PD"):
        # Mapeo de códigos de liga a hojas del Excel
        sheet_mapping = {
            "PL": "Inglaterra",
            "SA": "Italia",
            "PD": "Espa\u00f1a", # España
            "COL": "Colombia",
            "BL1": "Alemania",
            "FL1": "Francia",
            "PPL": "Portugal",
            "TSL": "Turquia",
            "DED": "Paises Bajos",
            "SPL": "Arabia Saudit",
            "J1": "Japon",
            "BEL": "Belgica",
            "CSL": "China",
            "MX": "Mexico ",
            "ELC": "Champio chip",
            "ECU": "Ecuador",
            "GRE": "Grecia"
        }
        
        base_dir = sys._MEIPASS if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))
        excel_path = os.path.join(base_dir, "TABLAS DE POSICIONES.xlsx")
        import openpyxl
        
        if os.path.exists(excel_path) and league in sheet_mapping:
            sheet_name = sheet_mapping[league]
            try:
                wb = openpyxl.load_workbook(excel_path, data_only=True)
                if sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    result = []
                    for r_idx, row in enumerate(sheet.iter_rows(values_only=False)):
                        # Comprobar si el primer elemento de la fila es una posición numérica
                        col0 = row[0].value
                        if col0 is not None:
                            try:
                                # Limpiar y parsear posición
                                pos_str = str(col0).strip().split('.')[0]
                                pos = int(pos_str)
                            except ValueError:
                                continue # No es una fila de equipo
                            
                            # Es una fila de equipo
                            team_cell = row[1]
                            team_name = str(team_cell.value).strip() if team_cell.value is not None else ""
                            if not team_name:
                                continue
                                
                            # Extraer URL de ESPN si existe
                            espn_url = team_cell.hyperlink.target if team_cell.hyperlink else ""
                            
                            # Parsear métricas numéricas con fallbacks seguros
                            def parse_int(cell_val, default=0):
                                if cell_val is None:
                                    return default
                                try:
                                    return int(float(str(cell_val).strip()))
                                except Exception:
                                    return default
                            
                            played = parse_int(row[2].value, 24)
                            won = parse_int(row[3].value, 0)
                            draw = parse_int(row[4].value, 0)
                            lost = parse_int(row[5].value, 0)
                            gd = parse_int(row[6].value, 0)
                            pts = parse_int(row[7].value, 0)
                            
                            result.append({
                                "position": pos,
                                "team": {
                                    "name": team_name,
                                    "crest": self.get_crest(team_name),
                                    "espn_url": espn_url
                                },
                                "playedGames": played,
                                "won": won,
                                "draw": draw,
                                "lost": lost,
                                "goalsFor": 0, # Opcionales
                                "goalsAgainst": 0,
                                "goalDifference": gd,
                                "points": pts
                            })
                    if result:
                        return sorted(result, key=lambda x: x["position"])
            except Exception as e:
                print(f"Error parsing Excel for standings: {e}")
                
        # FALLBACK: Simulación original si no existe el Excel o falla
        data = self._get_complete_table_data()
        teams = data.get(league, data["PD"])
        base_pts = 60
        result = []
        for i, t in enumerate(teams):
            won = max(0, (base_pts - i*2) // 3)
            draw = max(0, (base_pts - i*2) % 3)
            lost = max(0, 24 - won - draw)
            result.append({
                "position": i+1,
                "team": {"name": t, "crest": self.get_crest(t), "espn_url": ""},
                "playedGames": 24,
                "won": won, "draw": draw, "lost": lost,
                "points": max(5, base_pts - i*2),
                "goalsFor": max(10, 50-i),
                "goalsAgainst": max(8, 20+i),
                "goalDifference": max(-20, 30-i*2)
            })
        return result

    def get_fixtures(self, league="PD"):
        data = self._get_complete_table_data()
        teams = data.get(league, data["PD"])
        fixtures = []
        today = datetime.now()
        for i in range(0, min(12, len(teams)-1), 2):
            home_t = teams[i]
            away_t = teams[i+1]
            fixtures.append({
                "homeTeam": {"name": home_t, "id": self.master_teams.get(home_t, random.randint(1000, 9999))},
                "awayTeam": {"name": away_t, "id": self.master_teams.get(away_t, random.randint(1000, 9999))},
                "utcDate": (today + timedelta(days=random.randint(1, 7))).isoformat() + "Z"
            })
        return fixtures

    def get_team_stats(self, team_name):
        # ── Check real WC2026 profile first ──────────────────────────────────
        wc_profile = self.WC2026_TEAM_PROFILES.get(team_name)
        if wc_profile:
            p = wc_profile
            form_str = p["form"]
            form_score = sum(1.0 if r=="W" else 0.5 if r=="D" else 0.0 for r in form_str) / len(form_str)
            rank_factor = max(0.80, min(1.20, 1.0 + (50 - p["fifa_rank"]) * 0.004))
            recent_results = []
            comps = ["Clasificatoria", "Clasificatoria", "Amistoso", "Clasificatoria", "Amistoso"]
            rivals = ["Rival Zona", "Rival Zona", "Selección Top", "Rival Zona", "Sel. Amistoso"]
            score_map = {"W": "2-1", "D": "1-1", "L": "0-1"}
            for m_idx, result in enumerate(form_str):
                score_str = score_map[result]
                h_g, a_g = map(int, score_str.split("-"))
                g_scored = h_g
                g_conceded = a_g
                match_stat = self.generate_match_stats(g_scored, g_conceded, p["sofa_rating"])
                recent_results.append({
                    "competition": comps[m_idx],
                    "date": "Mar 2026", "time": "20:00",
                    "home": team_name, "away": rivals[m_idx],
                    "score": score_str, "result": result,
                    "stats": match_stat
                })
            
            hist_avg = self.calculate_averages_from_history(recent_results, team_name)
            
            return {
                "name": team_name,
                "crest": self.get_crest(team_name),
                "sofa_rating": p["sofa_rating"],
                "form": form_str,
                "form_score": round(form_score, 2),
                "rank_factor": round(rank_factor, 3),
                "fifa_rank": p["fifa_rank"],
                "confederation": p["confederation"],
                "key_player": p["key_player"],
                "wc_history": p["wc_history"],
                "style": p["style"],
                "summary": {
                    "goals_scored": round(hist_avg["avg_goals"] * 10),
                    "clean_sheets": max(1, round((1.5 - hist_avg["avg_goals_conceded"]) * 5)),
                    "avg_possession": p["possession"],
                    "avg_sofacore_rating": p["sofa_rating"],
                    "yellow_cards": round(hist_avg["avg_yellow_cards"] * 10),
                    "yellow_cards_per_game": hist_avg["avg_yellow_cards"],
                    "avg_goal_kicks": hist_avg["avg_goal_kicks"],
                    "avg_throw_ins": hist_avg["avg_throw_ins"],
                    "avg_saves": hist_avg["avg_saves"],
                },
                "attack": {
                    "goals_per_game": hist_avg["avg_goals"],
                    "total_shots": hist_avg["avg_total_shots"],
                    "shots_on_target": hist_avg["avg_shots_on_target"],
                    "big_chances_created": round(hist_avg["avg_shots_on_target"] * 0.4),
                    "key_passes": round(hist_avg["avg_shots_on_target"] * 2.0),
                    "corners": hist_avg["avg_corners"],
                },
                "defense": {
                    "goals_conceded": hist_avg["avg_goals_conceded"],
                    "tackles": round(14 + (1 - form_score) * 4),
                    "interceptions": round(9 + form_score * 3),
                    "clearances": round(14 + hist_avg["avg_goals_conceded"] * 4),
                    "duels_won": f"{round(48 + (p['sofa_rating'] - 6.5) * 3)}%",
                },
                "matches": {
                    "results": recent_results,
                    "fixtures": [
                        {"competition": "FIFA World Cup 2026", "date": "Jun 2026", "time": "18:00",
                         "home": team_name, "away": "Por definir", "score": "-"},
                    ]
                },
                "top_players": [
                    {"name": p["key_player"].split("–")[0].strip(), "rating": round(p["sofa_rating"]+0.6, 2),
                     "goals": round(p["goals_per_game"] * 8), "image": "https://api.sofascore.app/api/v1/player/895311/image"},
                    {"name": f"Portero de {team_name}", "rating": round(p["sofa_rating"]-0.1, 2),
                     "goals": 0, "image": "https://api.sofascore.app/api/v1/player/981504/image"},
                ],
                "info": {
                    "stadium": p["stadium"], "coach": p["coach"],
                    "country": team_name, "league": f"FIFA WC 2026 – {p['confederation']}",
                    "capacity": "50,000+",
                },
                "squad": [
                    {"name": f"Portero {team_name}", "pos": "G", "no": 1, "age": 28, "value": "€15M"},
                    {"name": f"Defensa {team_name}", "pos": "D", "no": 4, "age": 26, "value": "€35M"},
                    {"name": f"Centrocampista {team_name}", "pos": "M", "no": 10, "age": 25, "value": "€60M"},
                    {"name": p["key_player"].split("(")[0].strip(), "pos": "F", "no": 7, "age": 27, "value": "€90M"},
                ],
                "transfers": [],
            }

        # ── Fallback for club teams (random data) ──────────────────────────────
        info = self.team_database.get(team_name, {
            "stadium": f"Estadio de {team_name}",
            "coach": "Entrenador Pro",
            "country": "Internacional",
            "capacity": f"{random.randint(20, 80)},000"
        })
        if "capacity" not in info:
            info["capacity"] = f"{random.randint(20, 80)},000"
        rating = round(random.uniform(6.5, 7.8), 2)
        form = [random.choice(["W","D","L"]) for _ in range(5)]
        
        recent_results = [
            {"competition": "Champions League", "date": "18/02/26", "time": "21:00", "home": "Bayern Munich", "away": team_name, "score": "1-1", "result": "D"},
            {"competition": "Liga", "date": "14/02/26", "time": "21:00", "home": team_name, "away": "Athletic Club", "score": "2-1", "result": "W"},
            {"competition": "Copa", "date": "11/02/26", "time": "19:00", "home": "Atlético de Madrid", "away": team_name, "score": "0-1", "result": "W"},
            {"competition": "Champions League", "date": "04/02/26", "time": "21:00", "home": team_name, "away": "AC Milan", "score": "3-1", "result": "W"},
            {"competition": "Liga", "date": "01/02/26", "time": "21:00", "home": "FC Barcelona", "away": team_name, "score": "2-2", "result": "D"},
        ]
        
        for m in recent_results:
            try:
                h_g, a_g = map(int, m["score"].split("-"))
                g_scored = h_g if m["home"] == team_name else a_g
                g_conceded = a_g if m["home"] == team_name else h_g
            except:
                g_scored = random.randint(1, 2)
                g_conceded = random.randint(1, 2)
            m["stats"] = self.generate_match_stats(g_scored, g_conceded, rating)
            
        hist_avg = self.calculate_averages_from_history(recent_results, team_name)
        
        return {
            "name": team_name,
            "crest": self.get_crest(team_name),
            "sofa_rating": rating,
            "form": form,
            "form_score": sum(1.0 if r=="W" else 0.5 if r=="D" else 0.0 for r in form) / len(form),
            "rank_factor": 1.0,
            "fifa_rank": 50,
            "confederation": "UEFA",
            "key_player": f"Estrella de {team_name}",
            "wc_history": "Datos no disponibles",
            "style": "Estilo de juego competitivo",
            "summary": {
                "goals_scored": round(hist_avg["avg_goals"] * 10),
                "clean_sheets": max(1, round((1.5 - hist_avg["avg_goals_conceded"]) * 5)),
                "avg_possession": f"{random.randint(45,62)}%",
                "avg_sofacore_rating": rating,
                "yellow_cards": round(hist_avg["avg_yellow_cards"] * 10),
                "yellow_cards_per_game": hist_avg["avg_yellow_cards"],
                "avg_goal_kicks": hist_avg["avg_goal_kicks"],
                "avg_throw_ins": hist_avg["avg_throw_ins"],
                "avg_saves": hist_avg["avg_saves"],
            },
            "attack": {
                "goals_per_game": hist_avg["avg_goals"],
                "total_shots": hist_avg["avg_total_shots"],
                "shots_on_target": hist_avg["avg_shots_on_target"],
                "big_chances_created": round(hist_avg["avg_shots_on_target"] * 0.4),
                "key_passes": round(hist_avg["avg_shots_on_target"] * 2.0),
                "corners": hist_avg["avg_corners"],
            },
            "defense": {
                "goals_conceded": hist_avg["avg_goals_conceded"],
                "tackles": random.randint(12,18),
                "interceptions": random.randint(7,12),
                "clearances": round(14 + hist_avg["avg_goals_conceded"] * 4),
                "duels_won": f"{random.randint(48,58)}%"
            },
            "matches": {
                "results": recent_results,
                "fixtures": [
                    {"competition": "Champions League", "date": "04/03/26", "time": "21:00", "home": team_name, "away": "Bayern Munich", "score": "-"},
                    {"competition": "Liga", "date": "22/02/26", "time": "21:00", "home": "Rival D", "away": team_name, "score": "-"}
                ]
            },
            "top_players": [
                {"name": f"Estrella 1 de {team_name}", "rating": round(rating+0.5,2), "goals": random.randint(5,15), "image": "https://api.sofascore.app/api/v1/player/895311/image"},
                {"name": f"Estrella 2 de {team_name}", "rating": round(rating+0.3,2), "goals": random.randint(2,10), "image": "https://api.sofascore.app/api/v1/player/981504/image"}
            ],
            "info": info,
            "squad": [
                {"name": f"Portero {team_name}", "pos": "G", "no": 1, "age": 28, "value": "€10M"},
                {"name": f"Defensa {team_name}", "pos": "D", "no": 4, "age": 25, "value": "€30M"},
                {"name": f"Medio {team_name}",   "pos": "M", "no": 10, "age": 23, "value": "€50M"},
                {"name": f"Delantero {team_name}","pos": "F", "no": 7, "age": 24, "value": "€80M"}
            ],
            "transfers": [
                {"player": "Fichaje Invierno", "from": "Club Origen", "type": "Comprado", "fee": "€20M", "date": "Ene 2026"},
                {"player": "Venta Verano", "from": "Club Destino", "type": "Vendido", "fee": "€40M", "date": "Ago 2025"}
            ]
        }

    def generate_match_stats(self, goals_scored, goals_conceded, team_sofa_rating):
        # Generar estadísticas realistas para un partido específico basándose en goles e indicadores
        shots_on_target = max(1, goals_scored + random.randint(1, 4))
        total_shots = shots_on_target + random.randint(3, 10)
        
        # Las paradas defensivas aumentan con goles concedidos y azar
        saves = max(0, random.randint(1, 4) + (1 if goals_conceded == 0 else 0))
        
        corners = max(2, random.randint(2, 6) + int(team_sofa_rating - 6.0))
        yellow_cards = random.randint(1, 4)
        goal_kicks = max(3, random.randint(4, 10) + int(goals_conceded * 1.2))
        throw_ins = max(12, random.randint(15, 28) - int(team_sofa_rating - 7.0))
        
        return {
            "goals": goals_scored,
            "goals_conceded": goals_conceded,
            "throw_ins": throw_ins,
            "goal_kicks": goal_kicks,
            "yellow_cards": yellow_cards,
            "corners": corners,
            "shots_on_target": shots_on_target,
            "total_shots": total_shots,
            "saves": saves
        }

    def calculate_averages_from_history(self, results, team_name):
        if not results:
            return None
            
        sum_goals = 0
        sum_conceded = 0
        sum_throw_ins = 0
        sum_goal_kicks = 0
        sum_yellow_cards = 0
        sum_corners = 0
        sum_shots_on_target = 0
        sum_total_shots = 0
        sum_saves = 0
        
        count = len(results)
        for r in results:
            m_stats = r.get("stats", {})
            sum_goals += m_stats.get("goals", 1.5)
            sum_conceded += m_stats.get("goals_conceded", 1.0)
            sum_throw_ins += m_stats.get("throw_ins", 21)
            sum_goal_kicks += m_stats.get("goal_kicks", 7.5)
            sum_yellow_cards += m_stats.get("yellow_cards", 2.0)
            sum_corners += m_stats.get("corners", 5.0)
            sum_shots_on_target += m_stats.get("shots_on_target", 4.5)
            sum_total_shots += m_stats.get("total_shots", 10.0)
            sum_saves += m_stats.get("saves", 3.0)
            
        return {
            "avg_goals": round(sum_goals / count, 2),
            "avg_goals_conceded": round(sum_conceded / count, 2),
            "avg_throw_ins": round(sum_throw_ins / count, 1),
            "avg_goal_kicks": round(sum_goal_kicks / count, 1),
            "avg_yellow_cards": round(sum_yellow_cards / count, 1),
            "avg_corners": round(sum_corners / count, 1),
            "avg_shots_on_target": round(sum_shots_on_target / count, 1),
            "avg_total_shots": round(sum_total_shots / count, 1),
            "avg_saves": round(sum_saves / count, 1)
        }

    def get_teams_by_league(self, league):
        if league == "WORLD_CUP_2026":
            wc_data = self.get_worldcup_data()
            teams = []
            for g_name, g_data in wc_data.items():
                for t in g_data["teams"]:
                    teams.append(t["name"])
            return sorted(list(set(teams)))
        else:
            data = self._get_complete_table_data()
            return data.get(league, [])

    # ── WORLD CUP 2026 REAL TEAM PROFILES ────────────────────────────────────────
    # Based on FIFA Rankings (Nov 2025), CONMEBOL/UEFA/CONCACAF qualifying stats
    # and recent international match history
    WC2026_TEAM_PROFILES = {
        # ─── GROUP A ───
        "México": {
            "fifa_rank": 16, "confederation": "CONCACAF",
            "goals_per_game": 1.8, "goals_conceded": 1.0, "corners": 5.2, "shots_on_target": 4.8,
            "yellow_cards": 2.1, "possession": "52%",
            "form": ["W","W","D","W","L"],
            "style": "Contraataque veloz y disciplina defensiva",
            "key_player": "Hirving Lozano (PSV) – alero desequilibrante",
            "wc_history": "Siempre llega a octavos, maldición del quinto partido",
            "sofa_rating": 7.3, "stadium": "Estadio Azteca", "coach": "Javier Aguirre",
        },
        "Sudáfrica": {
            "fifa_rank": 67, "confederation": "CAF",
            "goals_per_game": 1.2, "goals_conceded": 1.3, "corners": 4.0, "shots_on_target": 3.5,
            "yellow_cards": 2.3, "possession": "44%",
            "form": ["D","W","L","D","W"],
            "style": "Físico y directo, fuerte en duelos",
            "key_player": "Percy Tau (Al Ahly) – creativo y veloz",
            "wc_history": "Sede 2010; debutante de vuelta en 2026",
            "sofa_rating": 6.7, "stadium": "FNB Stadium", "coach": "Hugo Broos",
        },
        "Corea del Sur": {
            "fifa_rank": 22, "confederation": "AFC",
            "goals_per_game": 1.6, "goals_conceded": 1.1, "corners": 5.0, "shots_on_target": 4.2,
            "yellow_cards": 1.8, "possession": "50%",
            "form": ["W","D","W","W","D"],
            "style": "Presión alta y transiciones rápidas",
            "key_player": "Son Heung-min (Tottenham) – goleador y capitán",
            "wc_history": "Semifinalista 2002; octavos 2022",
            "sofa_rating": 7.1, "stadium": "Seoul World Cup Stadium", "coach": "Hong Myung-bo",
        },
        "Chequia": {
            "fifa_rank": 40, "confederation": "UEFA",
            "goals_per_game": 1.4, "goals_conceded": 1.2, "corners": 4.8, "shots_on_target": 3.9,
            "yellow_cards": 2.0, "possession": "48%",
            "form": ["D","W","D","L","W"],
            "style": "Organizado, juego posicional con pressing medio",
            "key_player": "Tomáš Souček (West Ham) – motor del medio",
            "wc_history": "Cuarto lugar como Checoslovaquia en 1962",
            "sofa_rating": 6.9, "stadium": "Doosan Arena", "coach": "Ivan Hašek",
        },
        # ─── GROUP B ───
        "Canadá": {
            "fifa_rank": 48, "confederation": "CONCACAF",
            "goals_per_game": 2.1, "goals_conceded": 0.9, "corners": 5.5, "shots_on_target": 5.1,
            "yellow_cards": 1.9, "possession": "53%",
            "form": ["W","W","W","D","W"],
            "style": "Presión alta con físico y velocidad por bandas",
            "key_player": "Alphonso Davies (Bayern) – lateral de clase mundial",
            "wc_history": "Primera participación en 1986; regresa en 2026 como anfitrión",
            "sofa_rating": 7.2, "stadium": "BMO Field / BC Place", "coach": "Jesse Marsch",
        },
        "Bosnia": {
            "fifa_rank": 65, "confederation": "UEFA",
            "goals_per_game": 1.5, "goals_conceded": 1.4, "corners": 4.5, "shots_on_target": 4.0,
            "yellow_cards": 2.2, "possession": "47%",
            "form": ["W","D","W","L","W"],
            "style": "Directo, con pivote fuerte y extremos creativos",
            "key_player": "Edin Džeko (Fenerbahçe) – veterano goleador",
            "wc_history": "Primera vez en cuartos europeos 2014; debut mundialista 2026",
            "sofa_rating": 6.8, "stadium": "Bilino Polje", "coach": "Sergej Barbarez",
        },
        "Qatar": {
            "fifa_rank": 58, "confederation": "AFC",
            "goals_per_game": 1.1, "goals_conceded": 1.8, "corners": 4.2, "shots_on_target": 3.4,
            "yellow_cards": 2.0, "possession": "48%",
            "form": ["L","D","L","W","D"],
            "style": "Organizado defensivo con velocidad en transiciones",
            "key_player": "Akram Afif (Al Sadd) – goleador local",
            "wc_history": "Eliminado en grupo 2022 como anfitrión",
            "sofa_rating": 6.5, "stadium": "Lusail Stadium", "coach": "Marquez Lopez",
        },
        "Suiza": {
            "fifa_rank": 19, "confederation": "UEFA",
            "goals_per_game": 1.7, "goals_conceded": 0.9, "corners": 5.1, "shots_on_target": 4.5,
            "yellow_cards": 1.8, "possession": "54%",
            "form": ["W","W","D","W","W"],
            "style": "Sólido, organizado, difícil de batir",
            "key_player": "Granit Xhaka (Bayer Leverkusen) – motor del equipo",
            "wc_history": "Cuartos de final en 2022, eliminado por España",
            "sofa_rating": 7.2, "stadium": "St. Jakob-Park", "coach": "Murat Yakin",
        },
        # ─── GROUP C ───
        "Brasil": {
            "fifa_rank": 5, "confederation": "CONMEBOL",
            "goals_per_game": 2.4, "goals_conceded": 0.7, "corners": 6.2, "shots_on_target": 6.1,
            "yellow_cards": 1.9, "possession": "60%",
            "form": ["W","W","W","D","W"],
            "style": "Posesión + velocidad por bandas + alto pressing",
            "key_player": "Vinicius Jr. (Real Madrid) – mejor jugador del mundo 2024",
            "wc_history": "5 títulos (1958,62,70,94,2002), favorito eterno",
            "sofa_rating": 7.9, "stadium": "Maracanã", "coach": "Dorival Júnior",
        },
        "Marruecos": {
            "fifa_rank": 14, "confederation": "CAF",
            "goals_per_game": 1.6, "goals_conceded": 0.6, "corners": 5.0, "shots_on_target": 4.2,
            "yellow_cards": 2.1, "possession": "49%",
            "form": ["W","W","D","W","D"],
            "style": "Defensivamente sólido, con contragolpe mortal",
            "key_player": "Achraf Hakimi (PSG) – lateral derechocon proyección total",
            "wc_history": "Semifinalistas históricos en 2022",
            "sofa_rating": 7.4, "stadium": "Mohammed V", "coach": "Walid Regragui",
        },
        "Haití": {
            "fifa_rank": 82, "confederation": "CONCACAF",
            "goals_per_game": 1.0, "goals_conceded": 1.7, "corners": 3.8, "shots_on_target": 3.1,
            "yellow_cards": 2.4, "possession": "41%",
            "form": ["L","D","W","L","D"],
            "style": "Físico y defensivo, sorpresa del grupo",
            "key_player": "Frantzdy Pierrot – goleador de la clasificatoria",
            "wc_history": "Única aparición en 1974, sensación del grupo",
            "sofa_rating": 6.3, "stadium": "Estadio Sylvio Cator", "coach": "Marc Collat",
        },
        "Escocia": {
            "fifa_rank": 39, "confederation": "UEFA",
            "goals_per_game": 1.5, "goals_conceded": 1.1, "corners": 4.9, "shots_on_target": 4.0,
            "yellow_cards": 2.0, "possession": "50%",
            "form": ["W","D","W","D","W"],
            "style": "Intenso y vertical, presión alta",
            "key_player": "Andrew Robertson (Liverpool) – lateral de clase mundial",
            "wc_history": "Más de 20 años sin clasificar a un Mundial",
            "sofa_rating": 7.0, "stadium": "Hampden Park", "coach": "Steve Clarke",
        },
        # ─── GROUP D ───
        "Estados Unidos": {
            "fifa_rank": 11, "confederation": "CONCACAF",
            "goals_per_game": 1.9, "goals_conceded": 1.0, "corners": 5.3, "shots_on_target": 5.0,
            "yellow_cards": 1.8, "possession": "54%",
            "form": ["W","W","D","W","W"],
            "style": "Atletismo + pressing europeo moderno",
            "key_player": "Christian Pulisic (AC Milan) – capitán y goleador",
            "wc_history": "Anfitrión 2026; tercer lugar 1930",
            "sofa_rating": 7.4, "stadium": "MetLife Stadium", "coach": "Mauricio Pochettino",
        },
        "Paraguay": {
            "fifa_rank": 53, "confederation": "CONMEBOL",
            "goals_per_game": 1.3, "goals_conceded": 1.3, "corners": 4.5, "shots_on_target": 3.8,
            "yellow_cards": 2.3, "possession": "45%",
            "form": ["D","W","D","L","W"],
            "style": "Físico, agresivo y ordenado defensivamente",
            "key_player": "Miguel Almirón (Newcastle) – desequilibrante en mediapunta",
            "wc_history": "Cuartos de final 2010",
            "sofa_rating": 6.8, "stadium": "Estadio Defensores del Chaco", "coach": "Gustavo Alfaro",
        },
        "Australia": {
            "fifa_rank": 24, "confederation": "AFC",
            "goals_per_game": 1.5, "goals_conceded": 1.2, "corners": 4.8, "shots_on_target": 4.1,
            "yellow_cards": 1.9, "possession": "49%",
            "form": ["W","D","W","D","W"],
            "style": "Directo y físico, con buen bloque defensivo",
            "key_player": "Mathew Ryan (Real Sociedad) – portero clave",
            "wc_history": "Cuartos de final 2006, octavos 2022",
            "sofa_rating": 7.0, "stadium": "Stadium Australia", "coach": "Tony Popovic",
        },
        "Turquía": {
            "fifa_rank": 37, "confederation": "UEFA",
            "goals_per_game": 1.6, "goals_conceded": 1.1, "corners": 5.0, "shots_on_target": 4.3,
            "yellow_cards": 2.2, "possession": "51%",
            "form": ["W","W","D","W","D"],
            "style": "Rápido, combinativo, fuerte anímicamente",
            "key_player": "Hakan Çalhanoğlu (Inter Milan) – corazón del equipo",
            "wc_history": "Tercer lugar 2002; regresa tras larga ausencia",
            "sofa_rating": 7.1, "stadium": "Atatürk Olimpiyat", "coach": "Vincenzo Montella",
        },
        # ─── GROUP E ───
        "Alemania": {
            "fifa_rank": 12, "confederation": "UEFA",
            "goals_per_game": 2.0, "goals_conceded": 0.9, "corners": 5.8, "shots_on_target": 5.5,
            "yellow_cards": 1.7, "possession": "57%",
            "form": ["W","W","W","D","W"],
            "style": "Posesión + presión alta + eficiencia goleadora",
            "key_player": "Florian Wirtz (Bayer Leverkusen) – nuevo talismán",
            "wc_history": "4 títulos (1954,74,90,2014), potencia histórica",
            "sofa_rating": 7.6, "stadium": "Allianz Arena", "coach": "Julian Nagelsmann",
        },
        "Curazao": {
            "fifa_rank": 78, "confederation": "CONCACAF",
            "goals_per_game": 1.1, "goals_conceded": 1.6, "corners": 3.9, "shots_on_target": 3.2,
            "yellow_cards": 2.3, "possession": "42%",
            "form": ["D","L","W","D","L"],
            "style": "Contrariado, físico y directo",
            "key_player": "Cuco Martina – veterano del equipo",
            "wc_history": "Primera vez en el Mundial",
            "sofa_rating": 6.3, "stadium": "Ergilio Hato Stadium", "coach": "Remko Bicentini",
        },
        "Costa de Marfil": {
            "fifa_rank": 59, "confederation": "CAF",
            "goals_per_game": 1.4, "goals_conceded": 1.3, "corners": 4.7, "shots_on_target": 3.9,
            "yellow_cards": 2.2, "possession": "47%",
            "form": ["W","D","W","L","D"],
            "style": "Físico y técnico, con jugadores de ligas top",
            "key_player": "Sébastien Haller (Borussia Dortmund) – goleador de cabeza",
            "wc_history": "Siempre cerca pero sin avanzar de grupos",
            "sofa_rating": 6.9, "stadium": "Félix Houphouet-Boigny", "coach": "Emerse Faé",
        },
        "Ecuador": {
            "fifa_rank": 43, "confederation": "CONMEBOL",
            "goals_per_game": 1.5, "goals_conceded": 1.1, "corners": 4.9, "shots_on_target": 4.0,
            "yellow_cards": 2.0, "possession": "48%",
            "form": ["W","D","D","W","W"],
            "style": "Sólido defensivo, vertical en ataque",
            "key_player": "Moisés Caicedo (Chelsea) – mejor mediocampista de CONMEBOL",
            "wc_history": "Octavos de final en 2006",
            "sofa_rating": 7.0, "stadium": "Estadio Rodrigo Paz", "coach": "Sébastien Beccacece",
        },
        # ─── GROUP F ───
        "Países Bajos": {
            "fifa_rank": 7, "confederation": "UEFA",
            "goals_per_game": 2.1, "goals_conceded": 0.8, "corners": 5.9, "shots_on_target": 5.6,
            "yellow_cards": 1.8, "possession": "58%",
            "form": ["W","W","W","D","W"],
            "style": "Fútbol total moderno, presión alta y recuperación veloz",
            "key_player": "Virgil van Dijk (Liverpool) – el mejor defensor del mundo",
            "wc_history": "Finalista 1974,78,2010; semifinal 2014",
            "sofa_rating": 7.8, "stadium": "Johan Cruyff ArenA", "coach": "Ronald Koeman",
        },
        "Japón": {
            "fifa_rank": 17, "confederation": "AFC",
            "goals_per_game": 1.8, "goals_conceded": 0.9, "corners": 5.2, "shots_on_target": 4.8,
            "yellow_cards": 1.5, "possession": "53%",
            "form": ["W","W","D","W","W"],
            "style": "Presión intensa, juego técnico y compacto",
            "key_player": "Takefusa Kubo (Real Sociedad) – habilidad y creatividad",
            "wc_history": "Octavos consecutivos; aspirantes a cuartos",
            "sofa_rating": 7.4, "stadium": "Saitama Stadium", "coach": "Hajime Moriyasu",
        },
        "Ucrania": {
            "fifa_rank": 21, "confederation": "UEFA",
            "goals_per_game": 1.7, "goals_conceded": 1.0, "corners": 5.0, "shots_on_target": 4.4,
            "yellow_cards": 1.9, "possession": "51%",
            "form": ["W","W","D","W","D"],
            "style": "Organizado, disciplinado y con buen bloque",
            "key_player": "Mykhailo Mudryk (Chelsea) – desequilibrante por banda",
            "wc_history": "Cuartos de final 2006, debut mundialista reciente",
            "sofa_rating": 7.2, "stadium": "Olimpiyskiy", "coach": "Serhiy Rebrov",
        },
        "Túnez": {
            "fifa_rank": 29, "confederation": "CAF",
            "goals_per_game": 1.3, "goals_conceded": 1.0, "corners": 4.5, "shots_on_target": 3.7,
            "yellow_cards": 2.1, "possession": "46%",
            "form": ["D","W","D","W","D"],
            "style": "Organizado y disciplinado, difícil de batir",
            "key_player": "Wahbi Khazri – veterano goleador",
            "wc_history": "Primera selección africana en empatar con Europa en 1978",
            "sofa_rating": 6.8, "stadium": "Stade de Rades", "coach": "Jalel Kadri",
        },
        # ─── GROUP G ───
        "Bélgica": {
            "fifa_rank": 3, "confederation": "UEFA",
            "goals_per_game": 2.0, "goals_conceded": 0.7, "corners": 5.7, "shots_on_target": 5.3,
            "yellow_cards": 1.7, "possession": "57%",
            "form": ["W","W","D","W","W"],
            "style": "Técnico-físico, con generación dorada todavía vigente",
            "key_player": "Kevin De Bruyne (Man City) – mejor mediocampista del mundo",
            "wc_history": "Tercer lugar 2018; 2022 decepcionó en grupos",
            "sofa_rating": 7.8, "stadium": "King Baudouin Stadium", "coach": "Domenico Tedesco",
        },
        "Egipto": {
            "fifa_rank": 33, "confederation": "CAF",
            "goals_per_game": 1.4, "goals_conceded": 1.0, "corners": 4.6, "shots_on_target": 3.8,
            "yellow_cards": 2.1, "possession": "48%",
            "form": ["W","D","W","D","W"],
            "style": "Ordenado, depende de Salah para crear",
            "key_player": "Mohamed Salah (Liverpool) – leyenda viviente",
            "wc_history": "Regresa al Mundial después de larga ausencia",
            "sofa_rating": 7.1, "stadium": "Cairo International", "coach": "Hossam El-Badry",
        },
        "Irán": {
            "fifa_rank": 20, "confederation": "AFC",
            "goals_per_game": 1.5, "goals_conceded": 0.9, "corners": 4.8, "shots_on_target": 4.0,
            "yellow_cards": 2.2, "possession": "49%",
            "form": ["W","D","W","W","D"],
            "style": "Disciplinado defensivo con contragolpe rápido",
            "key_player": "Mehdi Taremi (Inter Milan) – goleador prolífico",
            "wc_history": "Siempre clasifica, raramente avanza de grupos",
            "sofa_rating": 7.0, "stadium": "Azadi Stadium", "coach": "Amir Ghalenoei",
        },
        "Nueva Zelanda": {
            "fifa_rank": 96, "confederation": "OFC",
            "goals_per_game": 1.0, "goals_conceded": 1.8, "corners": 3.8, "shots_on_target": 3.0,
            "yellow_cards": 2.1, "possession": "42%",
            "form": ["D","L","W","L","D"],
            "style": "Directo y físico, difícil de batir en el Pacífico",
            "key_player": "Chris Wood (Nottingham Forest) – delantero fuerte",
            "wc_history": "Participó en 1982 y 2010, ambas veces sin avanzar",
            "sofa_rating": 6.2, "stadium": "Eden Park", "coach": "Darren Bazeley",
        },
        # ─── GROUP H ───
        "España": {
            "fifa_rank": 8, "confederation": "UEFA",
            "goals_per_game": 2.2, "goals_conceded": 0.6, "corners": 6.0, "shots_on_target": 5.8,
            "yellow_cards": 1.6, "possession": "63%",
            "form": ["W","W","W","W","D"],
            "style": "Posesión absoluta + tiki-taka moderno + alta intensidad",
            "key_player": "Lamine Yamal (FC Barcelona) – prodigio de 18 años",
            "wc_history": "Campeón 2010; favorito 2026 con nueva generación dorada",
            "sofa_rating": 7.9, "stadium": "Estadio Metropolitano", "coach": "Luis de la Fuente",
        },
        "Cabo Verde": {
            "fifa_rank": 72, "confederation": "CAF",
            "goals_per_game": 1.1, "goals_conceded": 1.5, "corners": 3.9, "shots_on_target": 3.3,
            "yellow_cards": 2.3, "possession": "43%",
            "form": ["D","W","L","D","W"],
            "style": "Compacto y físico, con jugadores de Portugal",
            "key_player": "Garry Rodrigues – mejor jugador histórico",
            "wc_history": "Primera vez en el Mundial",
            "sofa_rating": 6.4, "stadium": "Estádio Nacional", "coach": "Bubista",
        },
        "Arabia Saudí": {
            "fifa_rank": 55, "confederation": "AFC",
            "goals_per_game": 1.2, "goals_conceded": 1.4, "corners": 4.3, "shots_on_target": 3.6,
            "yellow_cards": 2.2, "possession": "46%",
            "form": ["L","D","W","W","D"],
            "style": "Organizado y físico, con inversión en jugadores extranjeros",
            "key_player": "Salem Al-Dawsari – el que venció a Argentina 2022",
            "wc_history": "Histórico triunfo sobre Argentina 2022",
            "sofa_rating": 6.6, "stadium": "King Fahd Stadium", "coach": "Roberto Mancini",
        },
        "Uruguay": {
            "fifa_rank": 18, "confederation": "CONMEBOL",
            "goals_per_game": 1.7, "goals_conceded": 0.9, "corners": 5.0, "shots_on_target": 4.6,
            "yellow_cards": 2.0, "possession": "52%",
            "form": ["W","D","W","W","D"],
            "style": "Garra charrúa + técnica europea + solidez defensiva",
            "key_player": "Federico Valverde (Real Madrid) – mejor mediocampista de CONMEBOL",
            "wc_history": "Campeón 1930 y 1950; siempre competitivo",
            "sofa_rating": 7.5, "stadium": "Estadio Centenario", "coach": "Marcelo Bielsa",
        },
        # ─── GROUP I ───
        "Francia": {
            "fifa_rank": 2, "confederation": "UEFA",
            "goals_per_game": 2.3, "goals_conceded": 0.7, "corners": 6.1, "shots_on_target": 6.0,
            "yellow_cards": 1.6, "possession": "59%",
            "form": ["W","W","W","W","D"],
            "style": "Físico-técnico con jugadores de talla mundial en cada posición",
            "key_player": "Kylian Mbappé (Real Madrid) – favorito Balón de Oro",
            "wc_history": "Campeón 1998 y 2018; finalista 2022",
            "sofa_rating": 8.0, "stadium": "Stade de France", "coach": "Didier Deschamps",
        },
        "Senegal": {
            "fifa_rank": 23, "confederation": "CAF",
            "goals_per_game": 1.6, "goals_conceded": 0.8, "corners": 5.0, "shots_on_target": 4.3,
            "yellow_cards": 1.9, "possession": "50%",
            "form": ["W","D","W","W","W"],
            "style": "Físico y técnico, campeones africanos",
            "key_player": "Sadio Mané (Al Nassr) – líder histórico",
            "wc_history": "Cuartos de final 2002; octavos 2022",
            "sofa_rating": 7.3, "stadium": "Stade Léopold Sédar Senghor", "coach": "Aliou Cissé",
        },
        "Bolivia": {
            "fifa_rank": 84, "confederation": "CONMEBOL",
            "goals_per_game": 0.9, "goals_conceded": 2.0, "corners": 3.7, "shots_on_target": 2.9,
            "yellow_cards": 2.5, "possession": "39%",
            "form": ["L","D","L","W","L"],
            "style": "Resiste en altitude pero sufre fuera de La Paz",
            "key_player": "Marcelo Moreno Martins – goleador histórico",
            "wc_history": "Última vez en 1994; no ha ganado un partido mundialista",
            "sofa_rating": 6.0, "stadium": "Estadio Hernando Siles", "coach": "Oscar Villegas",
        },
        "Noruega": {
            "fifa_rank": 27, "confederation": "UEFA",
            "goals_per_game": 2.0, "goals_conceded": 0.9, "corners": 5.3, "shots_on_target": 5.0,
            "yellow_cards": 1.8, "possession": "53%",
            "form": ["W","W","D","W","W"],
            "style": "Directo con Haaland como referencia total",
            "key_player": "Erling Haaland (Man City) – mejor delantero del mundo",
            "wc_history": "No clasificaba desde 1998; vuelve con Haaland",
            "sofa_rating": 7.6, "stadium": "Ullevaal Stadion", "coach": "Ståle Solbakken",
        },
        "Suecia": {
            "fifa_rank": 28, "confederation": "UEFA",
            "goals_per_game": 1.9, "goals_conceded": 1.0, "corners": 5.4, "shots_on_target": 4.9,
            "yellow_cards": 1.8, "possession": "54%",
            "form": ["W","W","L","W","D"],
            "style": "Ataque fluido y físico con transiciones dinámicas",
            "key_player": "Viktor Gyökeres (Sporting CP) – delantero estrella en racha goleadora",
            "wc_history": "Subcampeón 1958; tercer lugar 1950 y 1994",
            "sofa_rating": 7.5, "stadium": "Friends Arena", "coach": "Jon Dahl Tomasson",
        },
        # ─── GROUP J ───
        "Argentina": {
            "fifa_rank": 1, "confederation": "CONMEBOL",
            "goals_per_game": 2.3, "goals_conceded": 0.6, "corners": 5.9, "shots_on_target": 6.2,
            "yellow_cards": 1.8, "possession": "58%",
            "form": ["W","W","W","W","W"],
            "style": "Campeones del mundo en plena forma – invictos 36 partidos",
            "key_player": "Lionel Messi (Inter Miami) – el mejor de la historia",
            "wc_history": "Campeón 1978, 1986, 2022; candidato número uno",
            "sofa_rating": 8.2, "stadium": "Estadio Monumental", "coach": "Lionel Scaloni",
        },
        "Argelia": {
            "fifa_rank": 44, "confederation": "CAF",
            "goals_per_game": 1.4, "goals_conceded": 1.1, "corners": 4.7, "shots_on_target": 3.8,
            "yellow_cards": 2.1, "possession": "48%",
            "form": ["W","D","W","D","D"],
            "style": "Técnico y ordenado, con mentalidad ganadora",
            "key_player": "Riyad Mahrez (Al Ahli) – extremo creativo",
            "wc_history": "Octavos en 2014; segunda participación histórica",
            "sofa_rating": 6.9, "stadium": "Stade Nelson Mandela", "coach": "Djamel Belmadi",
        },
        "Austria": {
            "fifa_rank": 26, "confederation": "UEFA",
            "goals_per_game": 1.7, "goals_conceded": 1.0, "corners": 5.0, "shots_on_target": 4.4,
            "yellow_cards": 1.9, "possession": "53%",
            "form": ["W","D","W","W","D"],
            "style": "Presión alta con juego posicional al estilo bundesliga",
            "key_player": "Marcel Sabitzer (Dortmund) – motor creativo",
            "wc_history": "Tercer lugar 1954; regresa tras décadas de ausencia",
            "sofa_rating": 7.1, "stadium": "Ernst-Happel-Stadion", "coach": "Ralf Rangnick",
        },
        "Jordania": {
            "fifa_rank": 71, "confederation": "AFC",
            "goals_per_game": 1.1, "goals_conceded": 1.5, "corners": 4.0, "shots_on_target": 3.3,
            "yellow_cards": 2.2, "possession": "44%",
            "form": ["D","D","W","L","D"],
            "style": "Organizado y compacto, difícil de batir en casa",
            "key_player": "Musa Al-Tamari – goleador de la zona asiática",
            "wc_history": "Primera vez en el Mundial",
            "sofa_rating": 6.4, "stadium": "Amman International", "coach": "Hossam Hassan",
        },
        # ─── GROUP K ───
        "Portugal": {
            "fifa_rank": 6, "confederation": "UEFA",
            "goals_per_game": 2.2, "goals_conceded": 0.7, "corners": 5.8, "shots_on_target": 5.7,
            "yellow_cards": 1.7, "possession": "58%",
            "form": ["W","W","W","D","W"],
            "style": "Técnico + físico + finalización clínica",
            "key_player": "Cristiano Ronaldo (Al Nassr) – leyenda con más goles en la historia",
            "wc_history": "Tercer lugar 1966; semifinal 2006; busca su primer título",
            "sofa_rating": 7.9, "stadium": "Estádio da Luz", "coach": "Roberto Martínez",
        },
        "Rep. D. Congo": {
            "fifa_rank": 56, "confederation": "CAF",
            "goals_per_game": 1.3, "goals_conceded": 1.2, "corners": 4.3, "shots_on_target": 3.7,
            "yellow_cards": 2.2, "possession": "46%",
            "form": ["W","D","W","D","L"],
            "style": "Físico, veloz por bandas, con buen pressing colectivo",
            "key_player": "Chancel Mbemba (Marsella) – defensa líder y capitán",
            "wc_history": "Participó en 1974 como Zaire; regresa históricamente al Mundial en 2026 tras pasar el repechaje intercontinental",
            "sofa_rating": 6.8, "stadium": "Stade des Martyrs", "coach": "Sébastien Desabre",
        },
        "Uzbekistán": {
            "fifa_rank": 62, "confederation": "AFC",
            "goals_per_game": 1.3, "goals_conceded": 1.3, "corners": 4.4, "shots_on_target": 3.7,
            "yellow_cards": 2.0, "possession": "46%",
            "form": ["W","D","W","D","W"],
            "style": "Físico y directo con buen pressing",
            "key_player": "Eldor Shomurodov (Cagliari) – goleador activo en Europa",
            "wc_history": "Primera vez en el Mundial",
            "sofa_rating": 6.6, "stadium": "Bunyodkor", "coach": "Srecko Katanec",
        },
        "Colombia": {
            "fifa_rank": 15, "confederation": "CONMEBOL",
            "goals_per_game": 1.9, "goals_conceded": 0.8, "corners": 5.4, "shots_on_target": 5.0,
            "yellow_cards": 1.9, "possession": "55%",
            "form": ["W","W","D","W","W"],
            "style": "Técnico y combinativo con jugadores creativos",
            "key_player": "Luis Díaz (Liverpool) – extremo desequilibrante",
            "wc_history": "Cuartos de final 2014; vuelve con gran generación",
            "sofa_rating": 7.6, "stadium": "Estadio El Campín", "coach": "Néstor Lorenzo",
        },
        # ─── GROUP L ───
        "Inglaterra": {
            "fifa_rank": 4, "confederation": "UEFA",
            "goals_per_game": 2.1, "goals_conceded": 0.7, "corners": 5.9, "shots_on_target": 5.6,
            "yellow_cards": 1.6, "possession": "60%",
            "form": ["W","W","D","W","W"],
            "style": "Presión alta + profundidad en plantilla + Southgate táctica sólida",
            "key_player": "Jude Bellingham (Real Madrid) – mejor jugador joven del mundo",
            "wc_history": "Campeón 1966; final 2020 Euros; finalista Euros 2024",
            "sofa_rating": 7.9, "stadium": "Wembley Stadium", "coach": "Gareth Southgate",
        },
        "Croacia": {
            "fifa_rank": 9, "confederation": "UEFA",
            "goals_per_game": 1.7, "goals_conceded": 0.9, "corners": 5.1, "shots_on_target": 4.7,
            "yellow_cards": 1.9, "possession": "54%",
            "form": ["D","W","W","D","W"],
            "style": "Técnico y resiliente, difícil de batir en torneos",
            "key_player": "Luka Modrić (Real Madrid) – cerebro del equipo y leyenda",
            "wc_history": "Finalista 2018; tercer lugar 2022",
            "sofa_rating": 7.6, "stadium": "Stadion Maksimir", "coach": "Zlatko Dalić",
        },
        "Ghana": {
            "fifa_rank": 60, "confederation": "CAF",
            "goals_per_game": 1.3, "goals_conceded": 1.4, "corners": 4.5, "shots_on_target": 3.8,
            "yellow_cards": 2.2, "possession": "46%",
            "form": ["W","D","L","W","D"],
            "style": "Físico y técnico, con jugadores en ligas europeas top",
            "key_player": "Mohammed Kudus (West Ham) – extremo creativo y goleador",
            "wc_history": "Cuartos de final 2010, penales ante Uruguay",
            "sofa_rating": 6.8, "stadium": "Baba Yara Sports Stadium", "coach": "Otto Addo",
        },
        "Panamá": {
            "fifa_rank": 74, "confederation": "CONCACAF",
            "goals_per_game": 1.0, "goals_conceded": 1.6, "corners": 4.0, "shots_on_target": 3.2,
            "yellow_cards": 2.4, "possession": "41%",
            "form": ["D","L","W","D","L"],
            "style": "Físico y defensivo, sin miedo al rival",
            "key_player": "Rolando Blackburn – goleador del equipo",
            "wc_history": "Segunda participación (debut 2018)",
            "sofa_rating": 6.3, "stadium": "Estadio Rommel Fernández", "coach": "Thomas Christiansen",
        },
    }

    def get_match_data_for_ai(self, h, v):
        return {"match": f"{h} vs {v}"}

    def get_apisports_stats(self, date="2026-06-22", match_query=None):
        if not self.apisports_key or "your_apisports" in self.apisports_key:
            return {"error": "API Key de API-Sports no configurada"}
        try:
            url = f"{self.apisports_url}/fixtures"
            params = {"date": date}
            response = requests.get(url, headers=self.apisports_headers, params=params)
            data = response.json()
            if not data.get("response"):
                return {"error": f"No se encontraron partidos para la fecha {date}"}
            fixture_id = None
            if match_query:
                for res in data["response"]:
                    home = res["teams"]["home"]["name"].lower()
                    away = res["teams"]["away"]["name"].lower()
                    if match_query.lower() in home or match_query.lower() in away:
                        fixture_id = res["fixture"]["id"]
                        break
            if not fixture_id:
                fixture_id = data["response"][0]["fixture"]["id"]
            stats_url = f"{self.apisports_url}/fixtures/statistics?fixture={fixture_id}"
            stats_response = requests.get(stats_url, headers=self.apisports_headers).json()
            return stats_response
        except Exception as e:
            return {"error": str(e)}
